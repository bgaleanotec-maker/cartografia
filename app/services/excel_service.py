"""
excel_service.py — Plantilla / importacion / exportacion de la tabla maestra SGI
=================================================================================
Lee y escribe el formato del Excel "Gestion - Apoyo Cartografia" usando openpyxl.
Maneja cabeceras agrupadas, desplegables (data validation) y parseo de fechas.
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.tracking import ESTADO_OPTIONS, STAGE_OPTIONS


# Definicion de columnas: (grupo, encabezado, llave logica)
COLUMNS = [
    ('Gestión Apoyo', 'Municipio', 'municipality'),
    ('Gestión Apoyo', 'EJECUTIVO COMERCIAL', 'commercial'),
    ('Gestión Apoyo', 'Relevancia', 'relevancia'),
    ('Gestión Apoyo', 'Proyecto', 'name'),
    ('Gestión Apoyo', 'Estado', 'stage'),
    ('Gestión Apoyo', 'Potencial SH', 'potential_clients'),
    ('Gestión Apoyo', 'Fecha asignación', 'assigned_date'),
    ('Gestión Apoyo', 'Fecha ejecución visita', 'visit_date'),
    ('Gestión Apoyo', 'Observaciones Apoyo Cartografía', 'support_notes'),
    ('Gestión Cartográfica', 'Numero de solicitud', 'cartografia.numero_solicitud'),
    ('Gestión Cartográfica', 'Fecha de solicitud', 'cartografia.fecha_solicitud'),
    ('Gestión Cartográfica', 'Fecha de respuesta', 'cartografia.fecha_respuesta'),
    ('Gestión Cartográfica', 'ESTADO', 'cartografia.estado'),
    ('Diseño de Redes', 'Numero de diseño', 'diseno.numero_diseno'),
    ('Diseño de Redes', 'Numero de solicitud', 'diseno.numero_solicitud'),
    ('Diseño de Redes', 'Fecha de solicitud', 'diseno.fecha_solicitud'),
    ('Diseño de Redes', 'Fecha de respuesta', 'diseno.fecha_respuesta'),
    ('Diseño de Redes', 'ESTADO', 'diseno.estado'),
    ('Informe Técnico', 'Numero de solicitud', 'informe_tecnico.numero_solicitud'),
    ('Informe Técnico', 'Fecha de solicitud', 'informe_tecnico.fecha_solicitud'),
    ('Informe Técnico', 'Requiere visita', 'informe_tecnico.requiere_visita'),
    ('Informe Técnico', 'Fecha visita cotización', 'informe_tecnico.fecha_visita_cotizacion'),
    ('Informe Técnico', 'Fecha de respuesta', 'informe_tecnico.fecha_respuesta'),
    ('Informe Técnico', 'ESTADO', 'informe_tecnico.estado'),
    ('Informe Técnico', 'Observaciones PDI', 'pdi_notes'),
]

GROUP_COLORS = {
    'Gestión Apoyo': '1E3A5F',
    'Gestión Cartográfica': '14532D',
    'Diseño de Redes': '4C1D95',
    'Informe Técnico': '7C2D12',
}

DATE_KEYS = {k for _, _, k in COLUMNS if 'fecha' in k or k in ('assigned_date', 'visit_date')}


def parse_date(value):
    """Parsea fecha desde Excel (datetime o string dd/mm/aa)."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _styled_header(ws):
    """Escribe la fila de grupo (1) y la fila de encabezados (2)."""
    thin = Side(style='thin', color='334155')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: grupos (celdas combinadas)
    col = 1
    groups = []
    for grp, _, _ in COLUMNS:
        if not groups or groups[-1][0] != grp:
            groups.append([grp, col, col])
        else:
            groups[-1][2] = col
        col += 1
    for grp, c1, c2 in groups:
        cell = ws.cell(row=1, column=c1, value=grp)
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor=GROUP_COLORS.get(grp, '374151'))

    # Fila 2: encabezados
    for idx, (grp, header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = Font(bold=True, color='E2E8F0', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor='1F2937')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 22
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = 'A3'


def build_template():
    """Genera la plantilla .xlsx vacia con cabeceras y desplegables."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cronograma visitas'
    _styled_header(ws)

    # Desplegables: ESTADO (3 columnas) y "Estado" (etapa) y Requiere visita
    dv_estado = DataValidation(type='list', formula1='"%s"' % ','.join(ESTADO_OPTIONS), allow_blank=True)
    dv_stage = DataValidation(type='list', formula1='"%s"' % ','.join([s.replace(',', ' ') for s in STAGE_OPTIONS]), allow_blank=True)
    dv_visita = DataValidation(type='list', formula1='"SI,NO"', allow_blank=True)
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_stage)
    ws.add_data_validation(dv_visita)

    for idx, (_, _, key) in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=2, column=idx).column_letter
        rng = f'{col_letter}3:{col_letter}500'
        if key.endswith('.estado'):
            dv_estado.add(rng)
        elif key == 'stage':
            dv_stage.add(rng)
        elif key == 'informe_tecnico.requiere_visita':
            dv_visita.add(rng)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def parse_import(file_stream):
    """Lee un .xlsx subido y devuelve lista de dicts (una por fila con datos)."""
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb['Cronograma visitas'] if 'Cronograma visitas' in wb.sheetnames else wb.worksheets[0]

    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or all(c is None or str(c).strip() == '' for c in r):
            continue
        record = {}
        for idx, (_, _, key) in enumerate(COLUMNS):
            value = r[idx] if idx < len(r) else None
            if value is not None and isinstance(value, str):
                value = value.strip()
            record[key] = value
        # Requiere proyecto y municipio para ser valido
        if not record.get('name'):
            continue
        rows.append(record)
    return rows


def build_export(projects):
    """Exporta proyectos (con sus process_requests) al formato del Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cronograma visitas'
    _styled_header(ws)

    row = 3
    for p in projects:
        procs = {pr.process_type: pr for pr in p.process_requests}
        values = _project_to_row(p, procs)
        for idx, (_, _, key) in enumerate(COLUMNS, start=1):
            ws.cell(row=row, column=idx, value=values.get(key, ''))
        row += 1

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _fmt(dt):
    return dt.strftime('%d/%m/%Y') if dt else ''


def _project_to_row(p, procs):
    cart = procs.get('cartografia')
    dis = procs.get('diseno')
    inf = procs.get('informe_tecnico')
    return {
        'municipality': p.municipality or '',
        'commercial': p.commercial.full_name if p.commercial else '',
        'relevancia': p.relevancia or '',
        'name': p.name or '',
        'stage': p.stage or '',
        'potential_clients': p.potential_clients or 0,
        'assigned_date': _fmt(p.assigned_date),
        'visit_date': _fmt(p.visit_date),
        'support_notes': p.support_notes or '',
        'cartografia.numero_solicitud': cart.numero_solicitud if cart else '',
        'cartografia.fecha_solicitud': _fmt(cart.fecha_solicitud) if cart else '',
        'cartografia.fecha_respuesta': _fmt(cart.fecha_respuesta) if cart else '',
        'cartografia.estado': cart.estado if cart else '',
        'diseno.numero_diseno': dis.numero_diseno if dis else '',
        'diseno.numero_solicitud': dis.numero_solicitud if dis else '',
        'diseno.fecha_solicitud': _fmt(dis.fecha_solicitud) if dis else '',
        'diseno.fecha_respuesta': _fmt(dis.fecha_respuesta) if dis else '',
        'diseno.estado': dis.estado if dis else '',
        'informe_tecnico.numero_solicitud': inf.numero_solicitud if inf else '',
        'informe_tecnico.fecha_solicitud': _fmt(inf.fecha_solicitud) if inf else '',
        'informe_tecnico.requiere_visita': ('SI' if inf.requiere_visita else 'NO') if inf else '',
        'informe_tecnico.fecha_visita_cotizacion': _fmt(inf.fecha_visita_cotizacion) if inf else '',
        'informe_tecnico.fecha_respuesta': _fmt(inf.fecha_respuesta) if inf else '',
        'informe_tecnico.estado': inf.estado if inf else '',
        'pdi_notes': p.pdi_notes or '',
    }
